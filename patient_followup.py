import io
import hashlib
from datetime import datetime

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

from export_tools import crear_visita_dict


VISITS_SHEET_PRIMARY = "VISITAS_SEGUIMINETO"
VISITS_SHEET_FALLBACK = "VISITAS_SEGUIMIENTO"
LEGACY_CURRENT_VISIT_SHEET = "VISITA ACTUAL"

REQUIRED_VISIT_COLUMNS = [
    "id_visita",
    "timestamp_registro",
    "fecha_visita",
    "nombre_paciente",
    "especie",
    "edad",
    "peso_kg",
    "bcs",
    "estado_corporal",
    "riesgo_nutricional",
    "etapa_vida",
    "condicion_fisiologica",
    "ajuste_senior",
    "objetivo_nutricional",
    "rer_kcal_dia",
    "mer_base_kcal_dia",
    "mer_final_kcal_dia",
    "alimento_evaluado",
    "gramos_dia_evaluados",
    "me_kcal_100g",
    "energia_aportada_kcal_dia",
    "cobertura_energia_pct",
    "gramos_recomendados_dia",
    "diferencia_gramos_dia",
    "cobertura_proteina_pct",
    "cobertura_grasa_pct",
    "resultado_nutricional",
    "decision_clinica",
    "diagnostico_nutricional",
    "recomendaciones",
    "observaciones",
    "profesional_responsable",
    "version_modelo",
    "fuente_dato",
    "permitir_edicion",
]

NUMERIC_COLUMNS = [
    "peso_kg",
    "bcs",
    "mer_final_kcal_dia",
    "energia_aportada_kcal_dia",
    "cobertura_energia_pct",
]


def _safe_float(value, default=0.0):
    try:
        if isinstance(value, str):
            value = value.replace(",", ".")
        return float(value)
    except Exception:
        return default


def _to_dataframe_rows(ws):
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    return pd.DataFrame(data_rows, columns=headers)


def _get_visits_sheet_name(sheet_names):
    if VISITS_SHEET_PRIMARY in sheet_names:
        return VISITS_SHEET_PRIMARY
    if VISITS_SHEET_FALLBACK in sheet_names:
        return VISITS_SHEET_FALLBACK
    return None


def _is_duplicate_row(df_visitas, candidate):
    candidate_id = str(candidate.get("id_visita", "")).strip()
    candidate_ts = str(candidate.get("timestamp_registro", "")).strip()
    if candidate_id and "id_visita" in df_visitas.columns:
        if df_visitas["id_visita"].astype(str).str.strip().eq(candidate_id).any():
            return True
    if candidate_ts and "timestamp_registro" in df_visitas.columns:
        if df_visitas["timestamp_registro"].astype(str).str.strip().eq(candidate_ts).any():
            return True
    return False


def _sort_visits_desc(df_visitas):
    df = df_visitas.copy()
    if "timestamp_registro" in df.columns:
        df["_sort_ts"] = pd.to_datetime(df["timestamp_registro"], errors="coerce")
    else:
        df["_sort_ts"] = pd.NaT
    if df["_sort_ts"].isna().all() and "fecha_visita" in df.columns:
        df["_sort_ts"] = pd.to_datetime(df["fecha_visita"], errors="coerce")
    return df.sort_values("_sort_ts", ascending=False, na_position="last").drop(columns=["_sort_ts"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_followup_workbook(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes))
    sheet_name = _get_visits_sheet_name(wb.sheetnames)
    if not sheet_name:
        return None, None, None, None, False, (
            f"No se encontró hoja de visitas. Use '{VISITS_SHEET_PRIMARY}' "
            f"o '{VISITS_SHEET_FALLBACK}'."
        )

    df_visitas = _to_dataframe_rows(wb[sheet_name])
    missing_cols = [c for c in REQUIRED_VISIT_COLUMNS if c not in df_visitas.columns]
    if missing_cols:
        return None, None, None, None, False, f"Columnas faltantes en {sheet_name}: {missing_cols}"

    migrated_rows = 0
    if LEGACY_CURRENT_VISIT_SHEET in wb.sheetnames:
        df_current = _to_dataframe_rows(wb[LEGACY_CURRENT_VISIT_SHEET])
        if not df_current.empty:
            for _, row in df_current.iterrows():
                candidate = {k: row.get(k, "") for k in REQUIRED_VISIT_COLUMNS}
                if not _is_duplicate_row(df_visitas, candidate):
                    df_visitas = pd.concat([df_visitas, pd.DataFrame([candidate])], ignore_index=True)
                    migrated_rows += 1

    df_visitas = _sort_visits_desc(df_visitas)
    metadatos_df = _to_dataframe_rows(wb["METADATOS"]) if "METADATOS" in wb.sheetnames else pd.DataFrame()
    return df_visitas, metadatos_df, sheet_name, migrated_rows, True, "OK"


def _prepare_plot_df(df_visitas):
    df = df_visitas.copy()
    df["fecha_plot"] = pd.to_datetime(df.get("fecha_visita"), errors="coerce")
    ts = pd.to_datetime(df.get("timestamp_registro"), errors="coerce")
    df["fecha_plot"] = df["fecha_plot"].where(df["fecha_plot"].notna(), ts)
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.sort_values("fecha_plot", ascending=True)


def _objective_indicator(ultima, peso_delta, bcs_delta):
    objetivo = str(ultima.get("objetivo_nutricional", "")).strip().lower()
    if not objetivo:
        return "⚪ Neutral"

    if any(x in objetivo for x in ["reduc", "bajar", "control"]):
        if peso_delta < -0.1 or bcs_delta < 0:
            return "🟢 Mejora"
        if abs(peso_delta) <= 0.1 and bcs_delta == 0:
            return "🟡 Estable"
        return "🔴 Empeora"

    if any(x in objetivo for x in ["recuper", "aument", "subir"]):
        if peso_delta > 0.1 or bcs_delta > 0:
            return "🟢 Mejora"
        if abs(peso_delta) <= 0.1 and bcs_delta == 0:
            return "🟡 Estable"
        return "🔴 Empeora"

    if "mant" in objetivo:
        if abs(peso_delta) <= 0.2 and bcs_delta == 0:
            return "🟢 Estable"
        return "🟡 Cambio observado"

    return "⚪ Neutral"


def _decision_clinica(df_plot):
    if len(df_plot) < 2:
        return "reevaluar"
    latest = df_plot.iloc[-1]
    prev = df_plot.iloc[-2]
    cobertura = _safe_float(latest.get("cobertura_energia_pct"))
    peso_delta = _safe_float(latest.get("peso_kg")) - _safe_float(prev.get("peso_kg"))
    bcs_delta = _safe_float(latest.get("bcs")) - _safe_float(prev.get("bcs"))

    if cobertura < 90 and (peso_delta < 0 or bcs_delta < 0):
        return "aumentar ración"
    if cobertura > 110 and (peso_delta > 0 or bcs_delta > 0):
        return "reducir ración"
    if 90 <= cobertura <= 110 and abs(peso_delta) <= 0.2 and bcs_delta == 0:
        return "mantener"
    return "reevaluar"


def _interpretacion_automatica(df_plot):
    if len(df_plot) < 2:
        return "No hay suficientes visitas para interpretar evolución."
    latest = df_plot.iloc[-1]
    prev = df_plot.iloc[-2]
    peso_delta = _safe_float(latest.get("peso_kg")) - _safe_float(prev.get("peso_kg"))
    bcs_delta = _safe_float(latest.get("bcs")) - _safe_float(prev.get("bcs"))
    cobertura = _safe_float(latest.get("cobertura_energia_pct"))

    msg = []
    if peso_delta > 0.2:
        msg.append("Se observa aumento de peso en la última visita.")
    elif peso_delta < -0.2:
        msg.append("Se observa disminución de peso en la última visita.")
    else:
        msg.append("El peso se mantiene estable entre visitas recientes.")

    if bcs_delta > 0:
        msg.append("El BCS presenta tendencia al alza.")
    elif bcs_delta < 0:
        msg.append("El BCS presenta tendencia a la baja.")
    else:
        msg.append("El BCS se mantiene estable.")

    if cobertura < 90:
        msg.append("La cobertura energética está por debajo del rango objetivo.")
    elif cobertura > 110:
        msg.append("La cobertura energética está por encima del rango objetivo.")
    else:
        msg.append("La cobertura energética está dentro del rango objetivo (90–110%).")

    return " ".join(msg)


def _build_current_visit_from_session():
    profile = st.session_state.get("profile", {})
    mascota = profile.get("mascota", {})

    nombre = st.session_state.get("nombre_mascota", mascota.get("nombre", "—")) or "—"
    especie = st.session_state.get("especie_mascota", mascota.get("especie", "perro"))
    edad = _safe_float(st.session_state.get("edad_mascota", mascota.get("edad", 0)))
    peso = _safe_float(st.session_state.get("peso_mascota", mascota.get("peso", 0)))
    bcs_raw = st.session_state.get("bcs_mascota", mascota.get("bcs", 5))
    bcs = int(round(_safe_float(bcs_raw, default=5)))
    bcs = max(1, min(9, bcs))

    mer_final = _safe_float(st.session_state.get("energia_actual", 0))
    rer = _safe_float(st.session_state.get("rer_actual", 0))
    mer_base = _safe_float(st.session_state.get("mer_base_actual", 0))
    senior_aplicado = bool(st.session_state.get("senior_aplicado_actual", False))
    etapa = st.session_state.get("etapa_mascota", mascota.get("etapa", "adulto"))
    condicion = st.session_state.get("condicion_mascota", mascota.get("condicion", "Castrado"))
    estado_corporal = st.session_state.get("estado_corporal_tab1", "—")
    riesgo = st.session_state.get("riesgo_nutricional_tab1", "—")
    diagnostico = st.session_state.get("interpretacion_diagnostico_tab1", "—")

    food_name = st.session_state.get("analysis_food_selector", "—")
    gramos = _safe_float(st.session_state.get(f"gramos_alimento_{food_name}", 0))
    me = _safe_float(st.session_state.get("me_alimento_actual", 0))
    energia = _safe_float(st.session_state.get("energia_aportada_actual", 0))
    cobertura = _safe_float(st.session_state.get("cobertura_energia_actual", 0))
    recomendados = _safe_float(st.session_state.get("gramos_recomendados_actual", 0))
    recomendaciones = st.session_state.get("recomendaciones_tab3", [])
    cob_pb = st.session_state.get("cobertura_proteina_actual")
    cob_ee = st.session_state.get("cobertura_grasa_actual")

    mascota_dict = {"nombre": nombre, "especie": especie}
    datos_energeticos = {
        "edad": edad,
        "peso": peso,
        "bcs": bcs,
        "estado_corporal": estado_corporal,
        "riesgo_nutricional": riesgo,
        "etapa": etapa,
        "condicion": condicion,
        "rer": rer,
        "mer_base": mer_base,
        "diagnostico": diagnostico,
    }
    datos_alimento = {
        "alimento": food_name,
        "gramos": gramos,
        "me": me,
        "aporte": energia,
        "cobertura": cobertura,
        "recomendados": recomendados,
    }
    return crear_visita_dict(
        mascota=mascota_dict,
        datos_energeticos=datos_energeticos,
        datos_alimento=datos_alimento,
        mer_final=mer_final,
        senior_applied=senior_aplicado,
        cob_pb=cob_pb,
        cob_ee=cob_ee,
        recomendaciones=recomendaciones,
    )


def _write_df_to_sheet(ws, df):
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in df.columns])


def _update_metadatos_sheet(ws_meta, df_visitas):
    now_iso = datetime.now().isoformat(timespec="seconds")
    ultima_fecha = ""
    if len(df_visitas) > 0:
        ultima_fecha = str(df_visitas.iloc[0].get("fecha_visita", ""))

    user_name = "Sistema UYWA"
    user = st.session_state.get("user")
    if isinstance(user, dict):
        user_name = user.get("name") or user.get("username") or user_name

    updates = {
        "ultima_visita_registrada": ultima_fecha,
        "numero_visitas": int(len(df_visitas)),
        "ultima_modificacion_por": user_name,
        "fecha_ultima_actualizacion": now_iso,
        "ultima_modificacion_fecha": now_iso,
    }

    key_col = 1
    value_col = 2
    existing_keys = {}
    for row_idx in range(2, ws_meta.max_row + 1):
        key = ws_meta.cell(row=row_idx, column=key_col).value
        if key:
            existing_keys[str(key).strip()] = row_idx

    for key, value in updates.items():
        if key in existing_keys:
            ws_meta.cell(row=existing_keys[key], column=value_col).value = value
        else:
            ws_meta.append([key, value])


def build_updated_workbook(file_bytes, df_visitas, visits_sheet_name):
    wb = load_workbook(io.BytesIO(file_bytes))
    if visits_sheet_name not in wb.sheetnames:
        wb.create_sheet(visits_sheet_name)
    _write_df_to_sheet(wb[visits_sheet_name], df_visitas[REQUIRED_VISIT_COLUMNS].copy())
    if "METADATOS" in wb.sheetnames:
        _update_metadatos_sheet(wb["METADATOS"], df_visitas)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def show_patient_followup():
    st.header("Seguimiento del paciente")
    uploaded_file = st.file_uploader(
        "Cargar ficha maestra (.xlsx)",
        type=["xlsx"],
        key="followup_master_file",
    )

    if not uploaded_file:
        st.info("Sube la ficha maestra descargada desde la pestaña Resumen y Exportar para iniciar el seguimiento.")
        return

    file_bytes = uploaded_file.getvalue()
    df_visitas, _, visits_sheet_name, migrated_rows, valid, message = load_followup_workbook(file_bytes)
    if not valid:
        st.error(message)
        return

    if migrated_rows:
        st.warning(f"Se migraron {migrated_rows} fila(s) desde '{LEGACY_CURRENT_VISIT_SHEET}'.")

    file_signature = f"{uploaded_file.name}:{len(file_bytes)}:{hashlib.md5(file_bytes).hexdigest()}"
    if "followup_df_visitas" not in st.session_state or st.session_state.get("followup_file_signature") != file_signature:
        st.session_state["followup_df_visitas"] = df_visitas.copy()
        st.session_state["followup_file_signature"] = file_signature
        st.session_state["followup_sheet_name"] = visits_sheet_name

    df_current = st.session_state["followup_df_visitas"].copy()
    df_current = _sort_visits_desc(df_current)
    st.session_state["followup_df_visitas"] = df_current

    st.subheader("Resumen rápido (última vs anterior)")
    if len(df_current) >= 2:
        ultima = df_current.iloc[0]
        penultima = df_current.iloc[1]
        peso_delta = _safe_float(ultima.get("peso_kg")) - _safe_float(penultima.get("peso_kg"))
        bcs_delta = _safe_float(ultima.get("bcs")) - _safe_float(penultima.get("bcs"))
        mer_delta = _safe_float(ultima.get("mer_final_kcal_dia")) - _safe_float(penultima.get("mer_final_kcal_dia"))
        cobertura_delta = _safe_float(ultima.get("cobertura_energia_pct")) - _safe_float(penultima.get("cobertura_energia_pct"))
        indicador = _objective_indicator(ultima, peso_delta, bcs_delta)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Peso (kg)", f"{_safe_float(ultima.get('peso_kg')):.2f}", f"{peso_delta:+.2f}")
        col2.metric("BCS", f"{_safe_float(ultima.get('bcs')):.1f}", f"{bcs_delta:+.1f}")
        col3.metric("MER final (kcal/día)", f"{_safe_float(ultima.get('mer_final_kcal_dia')):.1f}", f"{mer_delta:+.1f}")
        col4.metric("Cobertura energética (%)", f"{_safe_float(ultima.get('cobertura_energia_pct')):.1f}", f"{cobertura_delta:+.1f}")
        st.caption(f"Indicador clínico: {indicador}")
    else:
        st.info("Se requieren al menos 2 visitas para comparar evolución.")

    df_plot = _prepare_plot_df(df_current)
    st.subheader("Evolución clínica-nutricional")

    fig_peso = go.Figure()
    fig_peso.add_trace(go.Scatter(x=df_plot["fecha_plot"], y=df_plot["peso_kg"], mode="lines+markers", name="Peso (kg)"))
    fig_peso.update_layout(title="Peso vs tiempo", height=320, xaxis_title="Fecha", yaxis_title="Peso (kg)")

    fig_bcs = go.Figure()
    fig_bcs.add_trace(go.Scatter(x=df_plot["fecha_plot"], y=df_plot["bcs"], mode="lines+markers", name="BCS"))
    fig_bcs.update_layout(title="BCS vs tiempo", height=320, xaxis_title="Fecha", yaxis_title="BCS")

    fig_mer = go.Figure()
    fig_mer.add_trace(go.Scatter(x=df_plot["fecha_plot"], y=df_plot["mer_final_kcal_dia"], mode="lines+markers", name="MER final"))
    fig_mer.add_trace(go.Scatter(x=df_plot["fecha_plot"], y=df_plot["energia_aportada_kcal_dia"], mode="lines+markers", name="Energía aportada"))
    fig_mer.update_layout(title="MER final vs energía aportada", height=320, xaxis_title="Fecha", yaxis_title="kcal/día")

    fig_cov = go.Figure()
    fig_cov.add_trace(go.Scatter(x=df_plot["fecha_plot"], y=df_plot["cobertura_energia_pct"], mode="lines+markers", name="Cobertura (%)"))
    fig_cov.add_hline(y=100, line_dash="dash", line_color="gray")
    fig_cov.update_layout(title="Cobertura energética (%)", height=320, xaxis_title="Fecha", yaxis_title="%")

    g1, g2 = st.columns(2)
    with g1:
        st.plotly_chart(fig_peso, use_container_width=True)
        st.plotly_chart(fig_mer, use_container_width=True)
    with g2:
        st.plotly_chart(fig_bcs, use_container_width=True)
        st.plotly_chart(fig_cov, use_container_width=True)

    st.subheader("Interpretación automática")
    st.info(_interpretacion_automatica(df_plot))
    st.subheader("Decisión clínica sugerida")
    st.success(_decision_clinica(df_plot).capitalize())

    st.subheader("Histórico de visitas")
    show_all = st.checkbox("Mostrar todas las columnas", value=False, key="followup_show_all_cols")
    if show_all:
        st.dataframe(df_current, use_container_width=True, hide_index=True)
    else:
        key_cols = [
            "fecha_visita",
            "nombre_paciente",
            "peso_kg",
            "bcs",
            "mer_final_kcal_dia",
            "energia_aportada_kcal_dia",
            "cobertura_energia_pct",
            "decision_clinica",
            "diagnostico_nutricional",
        ]
        key_cols = [c for c in key_cols if c in df_current.columns]
        st.dataframe(df_current[key_cols], use_container_width=True, hide_index=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Agregar visita actual", use_container_width=True, key="followup_add_visit"):
            nueva_visita = _build_current_visit_from_session()
            if _is_duplicate_row(df_current, nueva_visita):
                st.warning("No se agregó la visita porque ya existe (id_visita o timestamp_registro).")
            else:
                df_current = pd.concat([pd.DataFrame([nueva_visita]), df_current], ignore_index=True)
                st.session_state["followup_df_visitas"] = _sort_visits_desc(df_current)
                st.success("Visita actual agregada al historial en memoria.")

    with c2:
        export_bytes = build_updated_workbook(
            file_bytes=file_bytes,
            df_visitas=st.session_state["followup_df_visitas"],
            visits_sheet_name=st.session_state.get("followup_sheet_name", visits_sheet_name),
        )
        st.download_button(
            "Descargar ficha actualizada",
            data=export_bytes,
            file_name=uploaded_file.name.replace(".xlsx", "_actualizada.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="followup_download_updated_workbook",
        )
