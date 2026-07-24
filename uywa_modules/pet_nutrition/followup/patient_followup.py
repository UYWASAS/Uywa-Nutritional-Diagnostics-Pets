import io
import hashlib
from datetime import datetime

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

from ..exports.export_tools import crear_visita_dict
from ..core.clinical_state import (
    clinical_state_is_ready,
    get_current_clinical_state,
    safe_float,
)


VISITS_SHEET_PRIMARY = "VISITAS_SEGUIMIENTO"
VISITS_SHEET_FALLBACK = "VISITAS_SEGUIMINETO"
LEGACY_CURRENT_VISIT_SHEET = "VISITA ACTUAL"

REQUIRED_VISIT_COLUMNS = [
    "id_visita",
    "timestamp_registro",
    "fecha_visita",
    "nombre_paciente",
    "especie",
    "edad",
    "peso_kg",
    "bcs",
    "estado_corporal",
    "riesgo_nutricional",
    "etapa_vida",
    "condicion_fisiologica",
    "ajuste_senior",
    "objetivo_nutricional",
    "rer_kcal_dia",
    "mer_base_kcal_dia",
    "mer_final_kcal_dia",
    "alimento_evaluado",
    "gramos_dia_evaluados",
    "me_kcal_100g",
    "energia_aportada_kcal_dia",
    "cobertura_energia_pct",
    "gramos_recomendados_dia",
    "diferencia_gramos_dia",
    "cobertura_proteina_pct",
    "cobertura_grasa_pct",
    "resultado_nutricional",
    "decision_clinica",
    "diagnostico_nutricional",
    "recomendaciones",
    "observaciones",
    "profesional_responsable",
    "version_modelo",
    "fuente_dato",
    "permitir_edicion",
]

NUMERIC_COLUMNS = [
    "peso_kg",
    "bcs",
    "mer_final_kcal_dia",
    "energia_aportada_kcal_dia",
    "cobertura_energia_pct",
]


def _to_dataframe_rows(ws):
    rows = list(ws.values)

    if not rows:
        return pd.DataFrame()

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    return pd.DataFrame(data_rows, columns=headers)


def _get_visits_sheet_name(sheet_names):
    if VISITS_SHEET_PRIMARY in sheet_names:
        return VISITS_SHEET_PRIMARY

    if VISITS_SHEET_FALLBACK in sheet_names:
        return VISITS_SHEET_FALLBACK

    return None


def _is_duplicate_row(df_visitas, candidate):
    candidate_id = str(candidate.get("id_visita", "")).strip()
    candidate_ts = str(candidate.get("timestamp_registro", "")).strip()

    if candidate_id and "id_visita" in df_visitas.columns:
        if df_visitas["id_visita"].astype(str).str.strip().eq(candidate_id).any():
            return True

    if candidate_ts and "timestamp_registro" in df_visitas.columns:
        if df_visitas["timestamp_registro"].astype(str).str.strip().eq(candidate_ts).any():
            return True

    return False


def _sort_visits_desc(df_visitas):
    df = df_visitas.copy()

    if "timestamp_registro" in df.columns:
        df["_sort_ts"] = pd.to_datetime(df["timestamp_registro"], errors="coerce")
    else:
        df["_sort_ts"] = pd.NaT

    if df["_sort_ts"].isna().all() and "fecha_visita" in df.columns:
        df["_sort_ts"] = pd.to_datetime(df["fecha_visita"], errors="coerce")

    return (
        df.sort_values("_sort_ts", ascending=False, na_position="last")
        .drop(columns=["_sort_ts"])
        .reset_index(drop=True)
    )


@st.cache_data(show_spinner=False)
def load_followup_workbook(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes))

    sheet_name = _get_visits_sheet_name(wb.sheetnames)

    if not sheet_name:
        return None, None, None, None, False, (
            f"No se encontró hoja de visitas. Use '{VISITS_SHEET_PRIMARY}'."
        )

    df_visitas = _to_dataframe_rows(wb[sheet_name])

    missing_cols = [c for c in REQUIRED_VISIT_COLUMNS if c not in df_visitas.columns]

    if missing_cols:
        return None, None, None, None, False, (
            f"Columnas faltantes en {sheet_name}: {missing_cols}"
        )

    migrated_rows = 0

    if LEGACY_CURRENT_VISIT_SHEET in wb.sheetnames:
        df_current = _to_dataframe_rows(wb[LEGACY_CURRENT_VISIT_SHEET])

        if not df_current.empty:
            for _, row in df_current.iterrows():
                candidate = {k: row.get(k, "") for k in REQUIRED_VISIT_COLUMNS}

                if not _is_duplicate_row(df_visitas, candidate):
                    df_visitas = pd.concat(
                        [df_visitas, pd.DataFrame([candidate])],
                        ignore_index=True,
                    )
                    migrated_rows += 1

    df_visitas = _sort_visits_desc(df_visitas)

    metadatos_df = (
        _to_dataframe_rows(wb["METADATOS"])
        if "METADATOS" in wb.sheetnames
        else pd.DataFrame()
    )

    return df_visitas, metadatos_df, sheet_name, migrated_rows, True, "OK"


def _prepare_plot_df(df_visitas):
    df = df_visitas.copy()

    df["fecha_plot"] = pd.to_datetime(df.get("fecha_visita"), errors="coerce")
    ts = pd.to_datetime(df.get("timestamp_registro"), errors="coerce")

    df["fecha_plot"] = df["fecha_plot"].where(df["fecha_plot"].notna(), ts)

    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df.sort_values("fecha_plot", ascending=True)


def _objective_indicator(ultima, peso_delta, bcs_delta):
    objetivo = str(ultima.get("objetivo_nutricional", "")).strip().lower()

    if not objetivo:
        return "⚪ Neutral"

    if any(x in objetivo for x in ["reduc", "bajar", "control"]):
        if peso_delta < -0.1 or bcs_delta < 0:
            return "🟢 Mejora"
        if abs(peso_delta) <= 0.1 and bcs_delta == 0:
            return "🟡 Estable"
        return "🔴 Empeora"

    if any(x in objetivo for x in ["recuper", "aument", "subir"]):
        if peso_delta > 0.1 or bcs_delta > 0:
            return "🟢 Mejora"
        if abs(peso_delta) <= 0.1 and bcs_delta == 0:
            return "🟡 Estable"
        return "🔴 Empeora"

    if "mant" in objetivo:
        if abs(peso_delta) <= 0.2 and bcs_delta == 0:
            return "🟢 Estable"
        return "🟡 Cambio observado"

    return "⚪ Neutral"


def _decision_clinica(df_plot):
    if len(df_plot) < 2:
        return "reevaluar"

    latest = df_plot.iloc[-1]
    prev = df_plot.iloc[-2]

    cobertura = safe_float(latest.get("cobertura_energia_pct"))
    peso_delta = safe_float(latest.get("peso_kg")) - safe_float(prev.get("peso_kg"))
    bcs_delta = safe_float(latest.get("bcs")) - safe_float(prev.get("bcs"))

    if cobertura < 90 and (peso_delta < 0 or bcs_delta < 0):
        return "aumentar ración"

    if cobertura > 110 and (peso_delta > 0 or bcs_delta > 0):
        return "reducir ración"

    if 90 <= cobertura <= 110 and abs(peso_delta) <= 0.2 and bcs_delta == 0:
        return "mantener"

    return "reevaluar"


def _interpretacion_automatica(df_plot):
    if len(df_plot) < 2:
        return "No hay suficientes visitas para interpretar evolución."

    latest = df_plot.iloc[-1]
    prev = df_plot.iloc[-2]

    peso_delta = safe_float(latest.get("peso_kg")) - safe_float(prev.get("peso_kg"))
    bcs_delta = safe_float(latest.get("bcs")) - safe_float(prev.get("bcs"))
    cobertura = safe_float(latest.get("cobertura_energia_pct"))

    msg = []

    if peso_delta > 0.2:
        msg.append("Se observa aumento de peso en la última visita.")
    elif peso_delta < -0.2:
        msg.append("Se observa disminución de peso en la última visita.")
    else:
        msg.append("El peso se mantiene estable entre visitas recientes.")

    if bcs_delta > 0:
        msg.append("El BCS presenta tendencia al alza.")
    elif bcs_delta < 0:
        msg.append("El BCS presenta tendencia a la baja.")
    else:
        msg.append("El BCS se mantiene estable.")

    if cobertura < 90:
        msg.append("La cobertura energética está por debajo del rango objetivo.")
    elif cobertura > 110:
        msg.append("La cobertura energética está por encima del rango objetivo.")
    else:
        msg.append("La cobertura energética está dentro del rango objetivo (90–110%).")

    return " ".join(msg)


def _generar_recomendaciones_basicas(pet, energy, food):
    recomendaciones = []

    bcs = pet.get("bcs", 5)
    cobertura = food.get("cobertura_energia", 0)

    if bcs < 5:
        recomendaciones.append("Monitorear recuperación de condición corporal.")
    elif bcs > 5:
        recomendaciones.append("Monitorear reducción gradual de peso y BCS.")
    else:
        recomendaciones.append("Mantener control periódico de peso y condición corporal.")

    if cobertura < 90:
        recomendaciones.append("Evaluar aumento gradual de la ración.")
    elif cobertura > 110:
        recomendaciones.append("Evaluar reducción de ración o alimento menos denso.")
    else:
        recomendaciones.append("La cobertura energética se encuentra en rango objetivo.")

    if energy.get("senior_aplicado"):
        recomendaciones.append("Mantener vigilancia nutricional por condición senior.")

    return recomendaciones


def _build_current_visit_from_clinical_state():
    ready, message = clinical_state_is_ready(require_food=True)

    if not ready:
        raise ValueError(message)

    state = get_current_clinical_state()

    pet = state["pet"]
    energy = state["energy"]
    food = state["food"]

    mascota_dict = {
        "nombre": pet.get("nombre", "—"),
        "especie": pet.get("especie", "perro"),
    }

    datos_energeticos = {
        "edad": pet.get("edad", 0),
        "peso": pet.get("peso", 0),
        "bcs": pet.get("bcs", 5),
        "estado_corporal": energy.get("estado_corporal", "—"),
        "riesgo_nutricional": energy.get("riesgo_nutricional", "—"),
        "etapa": pet.get("etapa", "adulto"),
        "condicion": pet.get("condicion", "—"),
        "rer": energy.get("rer", 0),
        "mer_base": energy.get("mer_base", 0),
        "diagnostico": energy.get("diagnostico", "—"),
    }

    datos_alimento = {
        "alimento": food.get("alimento", "—"),
        "gramos": food.get("gramos", 0),
        "me": food.get("me", 0),
        "aporte": food.get("energia_aportada", 0),
        "cobertura": food.get("cobertura_energia", 0),
        "recomendados": food.get("gramos_recomendados", 0),
    }

    recomendaciones = _generar_recomendaciones_basicas(pet, energy, food)

    return crear_visita_dict(
        mascota=mascota_dict,
        datos_energeticos=datos_energeticos,
        datos_alimento=datos_alimento,
        mer_final=energy.get("mer_final", 0),
        senior_applied=energy.get("senior_aplicado", False),
        cob_pb=food.get("cobertura_proteina"),
        cob_ee=food.get("cobertura_grasa"),
        recomendaciones=recomendaciones,
    )


def _write_df_to_sheet(ws, df):
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    ws.append(list(df.columns))

    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in df.columns])


def _update_metadatos_sheet(ws_meta, df_visitas):
    now_iso = datetime.now().isoformat(timespec="seconds")

    ultima_fecha = ""

    if len(df_visitas) > 0:
        ultima_fecha = str(df_visitas.iloc[0].get("fecha_visita", ""))

    user_name = "Sistema UYWA"
    user = st.session_state.get("user")

    if isinstance(user, dict):
        user_name = user.get("name") or user.get("username") or user_name

    updates = {
        "ultima_visita_registrada": ultima_fecha,
        "numero_visitas": int(len(df_visitas)),
        "ultima_modificacion_por": user_name,
        "fecha_ultima_actualizacion": now_iso,
        "ultima_modificacion_fecha": now_iso,
    }

    key_col = 1
    value_col = 2
    existing_keys = {}

    for row_idx in range(2, ws_meta.max_row + 1):
        key = ws_meta.cell(row=row_idx, column=key_col).value

        if key:
            existing_keys[str(key).strip()] = row_idx

    for key, value in updates.items():
        if key in existing_keys:
            ws_meta.cell(row=existing_keys[key], column=value_col).value = value
        else:
            ws_meta.append([key, value])


def build_updated_workbook(file_bytes, df_visitas, visits_sheet_name):
    wb = load_workbook(io.BytesIO(file_bytes))

    if visits_sheet_name not in wb.sheetnames:
        wb.create_sheet(visits_sheet_name)

    df_to_write = df_visitas.copy()

    for col in REQUIRED_VISIT_COLUMNS:
        if col not in df_to_write.columns:
            df_to_write[col] = ""

    df_to_write = df_to_write[REQUIRED_VISIT_COLUMNS].copy()

    _write_df_to_sheet(wb[visits_sheet_name], df_to_write)

    if VISITS_SHEET_FALLBACK in wb.sheetnames and visits_sheet_name != VISITS_SHEET_FALLBACK:
        try:
            del wb[VISITS_SHEET_FALLBACK]
        except Exception:
            pass

    if "METADATOS" in wb.sheetnames:
        _update_metadatos_sheet(wb["METADATOS"], df_to_write)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def _render_quick_summary(df_current):
    st.subheader("Resumen rápido")

    if len(df_current) >= 2:
        ultima = df_current.iloc[0]
        penultima = df_current.iloc[1]

        peso_delta = safe_float(ultima.get("peso_kg")) - safe_float(penultima.get("peso_kg"))
        bcs_delta = safe_float(ultima.get("bcs")) - safe_float(penultima.get("bcs"))
        mer_delta = safe_float(ultima.get("mer_final_kcal_dia")) - safe_float(penultima.get("mer_final_kcal_dia"))
        cobertura_delta = (
            safe_float(ultima.get("cobertura_energia_pct"))
            - safe_float(penultima.get("cobertura_energia_pct"))
        )

        indicador = _objective_indicator(ultima, peso_delta, bcs_delta)

        col1, col2, col3, col4 = st.columns(4)

        col1.metric("Peso (kg)", f"{safe_float(ultima.get('peso_kg')):.2f}", f"{peso_delta:+.2f}")
        col2.metric("BCS", f"{safe_float(ultima.get('bcs')):.1f}", f"{bcs_delta:+.1f}")
        col3.metric("MER final", f"{safe_float(ultima.get('mer_final_kcal_dia')):.1f}", f"{mer_delta:+.1f}")
        col4.metric(
            "Cobertura energética",
            f"{safe_float(ultima.get('cobertura_energia_pct')):.1f}%",
            f"{cobertura_delta:+.1f}",
        )

        st.caption(f"Indicador clínico: {indicador}")

    else:
        st.info("Se requieren al menos 2 visitas para comparar evolución.")


def _render_evolution_charts(df_current):
    df_plot = _prepare_plot_df(df_current)

    st.subheader("Evolución clínica-nutricional")

    if df_plot["fecha_plot"].isna().all():
        st.warning("No hay fechas válidas para graficar la evolución.")
        return df_plot

    fig_peso = go.Figure()
    fig_peso.add_trace(
        go.Scatter(
            x=df_plot["fecha_plot"],
            y=df_plot["peso_kg"],
            mode="lines+markers",
            name="Peso (kg)",
        )
    )
    fig_peso.update_layout(
        title="Peso vs tiempo",
        height=380,
        xaxis_title="Fecha",
        yaxis_title="Peso (kg)",
    )

    fig_bcs = go.Figure()
    fig_bcs.add_trace(
        go.Scatter(
            x=df_plot["fecha_plot"],
            y=df_plot["bcs"],
            mode="lines+markers",
            name="BCS",
        )
    )
    fig_bcs.update_layout(
        title="BCS vs tiempo",
        height=320,
        xaxis_title="Fecha",
        yaxis_title="BCS",
    )

    fig_mer = go.Figure()
    fig_mer.add_trace(
        go.Scatter(
            x=df_plot["fecha_plot"],
            y=df_plot["mer_final_kcal_dia"],
            mode="lines+markers",
            name="MER final",
        )
    )
    fig_mer.add_trace(
        go.Scatter(
            x=df_plot["fecha_plot"],
            y=df_plot["energia_aportada_kcal_dia"],
            mode="lines+markers",
            name="Energía aportada",
        )
    )
    fig_mer.update_layout(
        title="MER final vs energía aportada",
        height=320,
        xaxis_title="Fecha",
        yaxis_title="kcal/día",
    )

    fig_cov = go.Figure()
    fig_cov.add_trace(
        go.Scatter(
            x=df_plot["fecha_plot"],
            y=df_plot["cobertura_energia_pct"],
            mode="lines+markers",
            name="Cobertura (%)",
        )
    )
    fig_cov.add_hline(y=90, line_dash="dash", line_color="gray")
    fig_cov.add_hline(y=110, line_dash="dash", line_color="gray")
    fig_cov.update_layout(
        title="Cobertura energética (%)",
        height=320,
        xaxis_title="Fecha",
        yaxis_title="%",
    )

    g1, g2 = st.columns(2)

    with g1:
        st.plotly_chart(fig_peso, use_container_width=True)
        st.plotly_chart(fig_mer, use_container_width=True)

    with g2:
        st.plotly_chart(fig_bcs, use_container_width=True)
        st.plotly_chart(fig_cov, use_container_width=True)

    return df_plot


def _render_visit_table(df_current):
    st.subheader("Histórico de visitas")

    show_all = st.checkbox(
        "Mostrar todas las columnas",
        value=False,
        key="followup_show_all_cols",
    )

    if show_all:
        st.dataframe(df_current, use_container_width=True, hide_index=True)
    else:
        key_cols = [
            "fecha_visita",
            "nombre_paciente",
            "peso_kg",
            "bcs",
            "mer_final_kcal_dia",
            "energia_aportada_kcal_dia",
            "cobertura_energia_pct",
            "decision_clinica",
            "diagnostico_nutricional",
        ]

        key_cols = [c for c in key_cols if c in df_current.columns]

        st.dataframe(df_current[key_cols], use_container_width=True, hide_index=True)


def show_patient_followup():
    st.header("📈 Seguimiento Clínico-Nutricional del Paciente")

    uploaded_file = st.file_uploader(
        "Cargar ficha maestra (.xlsx)",
        type=["xlsx"],
        key="followup_master_file",
    )

    if not uploaded_file:
        st.info(
            "Sube la ficha maestra descargada desde la pestaña Resumen y Exportar "
            "para iniciar el seguimiento."
        )
        return

    file_bytes = uploaded_file.getvalue()

    df_visitas, _, visits_sheet_name, migrated_rows, valid, message = load_followup_workbook(file_bytes)

    if not valid:
        st.error(message)
        return

    if migrated_rows:
        st.warning(
            f"Se migraron {migrated_rows} fila(s) desde '{LEGACY_CURRENT_VISIT_SHEET}'."
        )

    if visits_sheet_name == VISITS_SHEET_FALLBACK:
        st.warning(
            "La ficha usa una hoja con nombre antiguo o mal escrito. "
            "Al descargar la ficha actualizada se normalizará a VISITAS_SEGUIMIENTO."
        )
        visits_sheet_name = VISITS_SHEET_PRIMARY

    file_signature = (
        f"{uploaded_file.name}:{len(file_bytes)}:{hashlib.md5(file_bytes).hexdigest()}"
    )

    if (
        "followup_df_visitas" not in st.session_state
        or st.session_state.get("followup_file_signature") != file_signature
    ):
        st.session_state["followup_df_visitas"] = df_visitas.copy()
        st.session_state["followup_file_signature"] = file_signature
        st.session_state["followup_sheet_name"] = visits_sheet_name

    df_current = st.session_state["followup_df_visitas"].copy()
    df_current = _sort_visits_desc(df_current)
    st.session_state["followup_df_visitas"] = df_current

    _render_quick_summary(df_current)

    df_plot = _render_evolution_charts(df_current)

    st.subheader("Interpretación automática")
    st.info(_interpretacion_automatica(df_plot))

    st.subheader("Decisión clínica sugerida")
    st.success(_decision_clinica(df_plot).capitalize())

    _render_visit_table(df_current)

    st.markdown("---")
    st.subheader("Actualizar ficha de seguimiento")

    ready, ready_message = clinical_state_is_ready(require_food=True)

    if ready:
        st.success("La visita actual está lista para agregarse al historial.")
    else:
        st.warning(f"No se puede agregar la visita actual: {ready_message}")

    c1, c2 = st.columns(2)

    with c1:
        if st.button(
            "Agregar visita actual",
            use_container_width=True,
            key="followup_add_visit",
            disabled=not ready,
        ):
            try:
                nueva_visita = _build_current_visit_from_clinical_state()

                if _is_duplicate_row(df_current, nueva_visita):
                    st.warning(
                        "No se agregó la visita porque ya existe "
                        "(id_visita o timestamp_registro)."
                    )
                else:
                    df_current = pd.concat(
                        [pd.DataFrame([nueva_visita]), df_current],
                        ignore_index=True,
                    )

                    st.session_state["followup_df_visitas"] = _sort_visits_desc(df_current)

                    st.success("Visita actual agregada al historial en memoria.")
                    st.rerun()

            except Exception as exc:
                st.error(f"No se pudo agregar la visita actual: {exc}")

    with c2:
        export_bytes = build_updated_workbook(
            file_bytes=file_bytes,
            df_visitas=st.session_state["followup_df_visitas"],
            visits_sheet_name=st.session_state.get(
                "followup_sheet_name",
                VISITS_SHEET_PRIMARY,
            ),
        )

        st.download_button(
            "Descargar ficha actualizada",
            data=export_bytes,
            file_name=uploaded_file.name.replace(".xlsx", "_actualizada.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="followup_download_updated_workbook",
        )
